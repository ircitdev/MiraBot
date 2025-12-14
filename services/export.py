"""
Export Service.
Экспорт статистики в CSV и Excel форматы.
"""

import csv
import io
from datetime import datetime, timedelta
from typing import List, Dict, Any
from loguru import logger

from database.repositories.user import UserRepository
from database.repositories.subscription import SubscriptionRepository
from database.repositories.conversation import ConversationRepository
from database.repositories.referral import ReferralRepository


class ExportService:
    """Сервис экспорта статистики."""

    def __init__(self):
        self.user_repo = UserRepository()
        self.subscription_repo = SubscriptionRepository()
        self.conversation_repo = ConversationRepository()
        self.referral_repo = ReferralRepository()

    async def export_summary_csv(self) -> io.BytesIO:
        """
        Экспорт общей статистики в CSV.

        Returns:
            BytesIO с CSV данными
        """
        now = datetime.now()
        today = now.replace(hour=0, minute=0, second=0, microsecond=0)
        week_ago = today - timedelta(days=7)
        month_ago = today - timedelta(days=30)

        # Собираем данные
        _, total_users = await self.user_repo.get_paginated(per_page=1)

        new_today = await self.user_repo.get_new_count(today)
        new_week = await self.user_repo.get_new_count(week_ago)
        new_month = await self.user_repo.get_new_count(month_ago)

        active_today = await self.user_repo.get_active_count(today)
        active_week = await self.user_repo.get_active_count(week_ago)

        premium_count = await self.subscription_repo.get_premium_count()
        free_count = await self.subscription_repo.get_free_count()

        total_referrals = await self.referral_repo.get_total_count()
        referrals_week = await self.referral_repo.get_count_since(week_ago)

        # Формируем CSV
        output = io.StringIO()
        writer = csv.writer(output)

        writer.writerow(["Метрика", "Значение"])
        writer.writerow(["Дата отчёта", now.strftime("%Y-%m-%d %H:%M")])
        writer.writerow([])
        writer.writerow(["Всего пользователей", total_users])
        writer.writerow([])
        writer.writerow(["Новые сегодня", new_today])
        writer.writerow(["Новые за неделю", new_week])
        writer.writerow(["Новые за месяц", new_month])
        writer.writerow([])
        writer.writerow(["Активные сегодня", active_today])
        writer.writerow(["Активные за неделю", active_week])
        writer.writerow([])
        writer.writerow(["Premium подписки", premium_count])
        writer.writerow(["Free подписки", free_count])
        writer.writerow(["Конверсия %", f"{premium_count / max(total_users, 1) * 100:.1f}"])
        writer.writerow([])
        writer.writerow(["Всего рефералов", total_referrals])
        writer.writerow(["Рефералов за неделю", referrals_week])

        # Конвертируем в bytes
        output.seek(0)
        bytes_io = io.BytesIO(output.getvalue().encode("utf-8-sig"))
        bytes_io.seek(0)

        logger.info("Exported summary statistics to CSV")
        return bytes_io

    async def export_users_csv(self) -> io.BytesIO:
        """
        Экспорт списка пользователей в CSV.

        Returns:
            BytesIO с CSV данными
        """
        # Получаем всех пользователей
        users, total = await self.user_repo.get_paginated(page=1, per_page=10000)

        output = io.StringIO()
        writer = csv.writer(output)

        # Заголовки
        writer.writerow([
            "Telegram ID",
            "Username",
            "Имя",
            "Персона",
            "Подписка",
            "Дата регистрации",
            "Последняя активность",
            "Сообщений всего",
            "Текст",
            "Голос",
            "Рефералов",
            "Заблокирован",
        ])

        # Данные
        for user in users:
            # Получаем подписку
            sub = await self.subscription_repo.get_active(user.id)
            plan = "Premium" if sub and sub.plan == "premium" else "Free"

            # Получаем статистику сообщений
            stats = await self.conversation_repo.get_user_message_stats(user.id)

            # Получаем рефералов
            referral_count = await self.referral_repo.count_by_referrer(user.id)

            writer.writerow([
                user.telegram_id,
                user.username or "",
                user.display_name or user.first_name or "",
                user.persona or "mira",
                plan,
                user.created_at.strftime("%Y-%m-%d %H:%M") if user.created_at else "",
                user.last_active_at.strftime("%Y-%m-%d %H:%M") if user.last_active_at else "",
                stats["total"],
                stats["text"],
                stats["voice"],
                referral_count,
                "Да" if user.is_blocked else "Нет",
            ])

        output.seek(0)
        bytes_io = io.BytesIO(output.getvalue().encode("utf-8-sig"))
        bytes_io.seek(0)

        logger.info(f"Exported {total} users to CSV")
        return bytes_io

    async def export_users_excel(self) -> io.BytesIO:
        """
        Экспорт списка пользователей в Excel.

        Returns:
            BytesIO с Excel данными
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        # Получаем всех пользователей
        users, total = await self.user_repo.get_paginated(page=1, per_page=10000)

        wb = Workbook()
        ws = wb.active
        ws.title = "Пользователи"

        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center")

        # Заголовки
        headers = [
            "Telegram ID",
            "Username",
            "Имя",
            "Персона",
            "Подписка",
            "Дата регистрации",
            "Последняя активность",
            "Сообщений всего",
            "Текст",
            "Голос",
            "Рефералов",
            "Заблокирован",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Данные
        for row, user in enumerate(users, 2):
            # Получаем подписку
            sub = await self.subscription_repo.get_active(user.id)
            plan = "Premium" if sub and sub.plan == "premium" else "Free"

            # Получаем статистику сообщений
            stats = await self.conversation_repo.get_user_message_stats(user.id)

            # Получаем рефералов
            referral_count = await self.referral_repo.count_by_referrer(user.id)

            ws.cell(row=row, column=1, value=user.telegram_id)
            ws.cell(row=row, column=2, value=user.username or "")
            ws.cell(row=row, column=3, value=user.display_name or user.first_name or "")
            ws.cell(row=row, column=4, value=user.persona or "mira")
            ws.cell(row=row, column=5, value=plan)
            ws.cell(row=row, column=6, value=user.created_at.strftime("%Y-%m-%d %H:%M") if user.created_at else "")
            ws.cell(row=row, column=7, value=user.last_active_at.strftime("%Y-%m-%d %H:%M") if user.last_active_at else "")
            ws.cell(row=row, column=8, value=stats["total"])
            ws.cell(row=row, column=9, value=stats["text"])
            ws.cell(row=row, column=10, value=stats["voice"])
            ws.cell(row=row, column=11, value=referral_count)
            ws.cell(row=row, column=12, value="Да" if user.is_blocked else "Нет")

        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Сохраняем в BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f"Exported {total} users to Excel")
        return output

    async def export_referrals_csv(self) -> io.BytesIO:
        """
        Экспорт списка рефералов в CSV.

        Returns:
            BytesIO с CSV данными
        """
        # Получаем топ рефереров
        top_referrers = await self.referral_repo.get_top_referrers(limit=1000)

        output = io.StringIO()
        writer = csv.writer(output)

        # Заголовки
        writer.writerow([
            "Место",
            "Telegram ID",
            "Имя",
            "Username",
            "Количество рефералов",
        ])

        # Данные
        for i, ref in enumerate(top_referrers, 1):
            writer.writerow([
                i,
                ref.get("telegram_id", ""),
                ref.get("display_name") or ref.get("first_name", ""),
                ref.get("username", ""),
                ref.get("referral_count", 0),
            ])

        output.seek(0)
        bytes_io = io.BytesIO(output.getvalue().encode("utf-8-sig"))
        bytes_io.seek(0)

        logger.info(f"Exported {len(top_referrers)} referrers to CSV")
        return bytes_io

    async def export_messages_stats_csv(self) -> io.BytesIO:
        """
        Экспорт статистики сообщений по дням.

        Returns:
            BytesIO с CSV данными
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Заголовки
        writer.writerow([
            "Дата",
            "Всего сообщений",
        ])

        # Получаем данные за последние 30 дней
        now = datetime.now()
        for days_ago in range(30):
            date = now - timedelta(days=days_ago)
            date_start = date.replace(hour=0, minute=0, second=0, microsecond=0)
            date_end = date_start + timedelta(days=1)

            count = await self.conversation_repo.get_messages_count_since(date_start)
            # Примечание: нужен метод для подсчёта за конкретный день

            writer.writerow([
                date_start.strftime("%Y-%m-%d"),
                count,
            ])

        output.seek(0)
        bytes_io = io.BytesIO(output.getvalue().encode("utf-8-sig"))
        bytes_io.seek(0)

        return bytes_io


# Глобальный экземпляр
export_service = ExportService()
